# import module
import openpyxl
import math
  
# load excel with its path
wrkbk = openpyxl.load_workbook("Fashion_Biz_AU_NZ_-_Colour_Hex.xlsx", data_only=True)

# Make sure the sheet name is same as on the excel
sh = wrkbk["Sheet1"]


myDict = {}
  
# iterate through excel and display data
for i in range(3, 180):
    print("\n")
    print("Row ", i, " data :")
      
    cell_hex = sh.cell(row=i, column=2).value       
    cell_col_name = sh.cell(row=i, column=1).value

    if cell_hex == "" or cell_hex == None:
        continue
    cell_hex = "#" + cell_hex
    myDict[cell_col_name.rstrip()] = cell_hex

print(myDict)

