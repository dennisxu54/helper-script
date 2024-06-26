# import module
import openpyxl
import math
  
# load excel with its path
wrkbk = openpyxl.load_workbook("Fashion Biz AU NZ - Product Details.xlsx", data_only=True)

# Make sure the sheet name is same as on the excel
sh = wrkbk["POLOS"]

list1 = []
myDict = {}
  
# iterate through excel and display data
for i in range(2, 205):
    print("\n")
    print("Row ", i, " data :")
      
    cell_product_slug = sh.cell(row=i, column=3).value
    if cell_product_slug == " " or cell_product_slug == None or cell_product_slug == "Style" or cell_product_slug == "STYLE":
        continue
    if cell_product_slug == "WORKS PANTS & SHORTS" or cell_product_slug == "ENGINEERED OUTERWEAR" or cell_product_slug == "OVERALLS" or cell_product_slug == "FIRE ARMOUR":
        continue
    if cell_product_slug == "POLOS" or cell_product_slug == "TEES" or cell_product_slug == "SHIRTS" or cell_product_slug == "BIZ SEPARATES" or cell_product_slug == "KNITWEAR":
        continue
    if cell_product_slug == "HOODIES & FLEECE" or cell_product_slug == "ACTIVEWEAR" or cell_product_slug == "JACKETS" or cell_product_slug == "HOSPITALITY" or cell_product_slug == "HEALTH & BEAUTY":
        continue
    if cell_product_slug == "HEALTHWEAR" or cell_product_slug == "CASUALWEAR" or cell_product_slug == "TOPS" or cell_product_slug == "SEPARATES":
        continue
    print(cell_product_slug)
    pro_slug = str(cell_product_slug).strip()
    pro_slug = pro_slug.lower()
    #key = pro_slug[0].lower()
    #key = str(cell_product_slug).lower() 
    #key = str(cell_product_slug)

    name = sh.cell(row=i, column=4).value
    size_range = sh.cell(row=i, column=6).value
    size_fit = sh.cell(row=i, column=7).value
    feature_icon = sh.cell(row=i, column=8).value
    description = sh.cell(row=i, column=9).value
    if description != None:
        description = description.replace('\n',"")
    fabric = sh.cell(row=i, column=10).value
    feature = sh.cell(row=i, column=11).value
    fabric_care = sh.cell(row=i, column=12).value
    size_model = sh.cell(row=i, column=13).value
    size_height = sh.cell(row=i, column=14).value
    sleeve_length = sh.cell(row=i, column=16).value
    fabric_type = sh.cell(row=i, column=17).value
    garment_type = sh.cell(row=i, column=18).value
    sub_brand = sh.cell(row=i, column=19).value
    feature_color = sh.cell(row=i, column=20).value
    #value = [cell_tier_price, cell_tier_price*2]
    #print(value)
    """
    if cell_tier_price1:
        features = cell_tier_price1.split(";")
    else:
        features = ""
    """
    
    new_list_value = [pro_slug, name, size_range, size_fit, feature_icon, description, fabric, feature, fabric_care, size_model, size_height, 
                      sleeve_length, fabric_type, garment_type, sub_brand, feature_color]
    list1.append(new_list_value)

#if you do want a unique list with no duplicates, remember to remove square brackets from above
#unique_list = list(dict.fromkeys(list1))
#print(unique_list)
print(list1)
