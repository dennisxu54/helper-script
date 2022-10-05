# import module
import openpyxl
import math
  
# load excel with its path
wrkbk = openpyxl.load_workbook("Copy of Biz Collection - Sizing.xlsx", data_only=True)

sh = wrkbk["Sheet1"]

size_names = ["XXS", "XS", "S", "M", "L", "XL", "2XL", "3XL", "4XL", "5XL", "6XL", "7XL", "2", "4", "6", "8", "10", "12", "14", "16", 
      "18", "20", "22", "24", "26", "28", "30", "72", "77", "82", "87", "92", "97", "102", "107", "112", "117", "122", "127", "132", "137", "142"]
list1 = []
myDict = {}
   
# iterate through excel and display data
for i in range(600, 980):
    print("\n")
    print("Row ", i, " data :")
      
    cell_product_size_name = sh.cell(row=i, column=2).value
    if cell_product_size_name == " " or cell_product_size_name == None or cell_product_size_name == "-" or cell_product_size_name == "":
        continue
    print(cell_product_size_name)
    key = str(cell_product_size_name)
    if "length" in key:
        key = key.replace("length", "Length")
    if "Lenth" in key:
        key = key.replace("Lenth", "Length")
    if "SLeeve" in key:
        key = key.replace("SLeeve", "Sleeve")
    if "  " in key:
        key = key.replace("  ", " ")
    if "from" in key:
        key = key.replace("from", "From")
    if "CNB" in key:
        key = key.replace("CNB", "CBN")
    if "\n" in key:
        key = key.replace("\n", "")
    if " ( CM)" in key:
        key = key.replace(" ( CM)", " (CM)")
    if "(CM)" not in key:
        key = key + " (CM)"
    if "  (CM)" in key:
        key = key.replace("  (CM)", " (CM)")
    if "1/2" in key:
        key = key.replace("1/2", "½")
    if "chest" in key:
        key = key.replace("chest", "Chest")
        
    if "Outleg From Waist (CM)" in key:
        key = "Outer Leg (cm)"
    if "Full Leg Length (CM)" in key:
        key = "Leg Length (cm)"
    if "Cut to Fit Sizes (CM)" in key:
        key = "To Fit Sizes"
    if "Inleg (CM)" in key or "Full Leg Length (CM)" in key:
        key = "Leg Length (cm)"
    if "Sleeve Length From CBN (CM)" in key:
        key = "Sleeve Length (cm)"
    if "Sleeve Length From HSP (CM)" in key:
        key = "Sleeve Length (cm)"
        
    if "(CM)" in key:
        key = key.replace("(CM)", "(cm)")


    size_list = []
    index = 0
    for x in range(3, 45):
        cell_tier_price1 = sh.cell(row=i, column=x).value
        if cell_tier_price1:
            index += 1
            size_list.append({"key": size_names[x-3], "value": str(cell_tier_price1), "position": index})
            print({"key": size_names[x-3], "value": str(cell_tier_price1), "position": index})
        else:
            continue

    cell_product_slug = sh.cell(row=i, column=1).value
    slug = str(cell_product_slug).lower()
    slug = slug.strip()
        
    myDict = {"size_details": size_list}
    list1.append([slug, key, myDict])



#if you do want a unique list with no duplicates
#unique_list = list(dict.fromkeys(list1))
#print(unique_list)
print(list1)
