# import module
import openpyxl
import math
  
# load excel with its path
wrkbk = openpyxl.load_workbook("All 2022 Biz Coll launch items.xlsx", data_only=True)

sh = wrkbk["Sheet1"]

list1 = []
myDict = {}
  
# iterate through excel and display data
for i in range(2, 920):
    print("\n")
    print("Row ", i, " data :")
      
    cell_product_slug = sh.cell(row=i, column=1).value
    if cell_product_slug == " " or cell_product_slug == None or cell_product_slug == "-":
        continue
    print(cell_product_slug)
    pro_slug = cell_product_slug.replace(" ", "")
    pro_slug = pro_slug.split("/")
    #key = pro_slug[0].lower()
    #key = str(cell_product_slug).lower() 
    #key = str(cell_product_slug)

    cell_tier_price1 = sh.cell(row=i, column=4).value
    cell_tier_color = sh.cell(row=i, column=8).value
    """
    cell_tier_price1 = cell_tier_price1.lower()
    cell_tier_price2 = sh.cell(row=i, column=8).value
    cell_tier_price3 = sh.cell(row=i, column=9).value
    
    if cell_tier_price == None:
        continue
    if type(cell_tier_price) == float:
        pass
    else:
        cell_tier_price = float(cell_tier_price[1:])
    """
    #value = [cell_tier_price, cell_tier_price*2]
    #print(value)
    """
    if cell_tier_price1:
        features = cell_tier_price1.split(";")
    else:
        features = ""
    """
    
    new_list_value = [pro_slug[0], pro_slug[2], cell_tier_color, cell_tier_price1]
    list1.append(new_list_value)

#if you do want a unique list with no duplicates, remember to remove square brackets from above
#unique_list = list(dict.fromkeys(list1))
#print(unique_list)
print(list1)
