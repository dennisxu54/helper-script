# import module
import openpyxl
import math
  
# load excel with its path
wrkbk = openpyxl.load_workbook("2024 Biz Corporates - Pricelist Reseller AU_September.xlsx", data_only=True)

# Make sure the sheet name is same as on the excel
sh = wrkbk["Biz Corporates AU"]

list1 = []
myDict = {}
  
# iterate through excel and display data
for i in range(7, 136):
    print("\n")
    print("Row ", i, " data :")
    category_name = [" ", "Style", "STYLE", "WORKS PANTS & SHORTS", "ENGINEERED OUTERWEAR", "OVERALLS", "FIRE ARMOUR", "POLOS", "TEES", 
                     "SHIRTS", "BIZ SEPARATES", "KNITWEAR", "HOODIES & FLEECE", "ACTIVEWEAR", "JACKETS", "HOSPITALITY", "HEALTH & BEAUTY",
                     "HEALTHWEAR", "CASUALWEAR", "TOPS", "SEPARATES", 'PAGE', 'SIENA SUITING', 'COOL STRETCH SUITING', 'COMFORT WOOL STRETCH SUITING',
                     'SMART CASUAL SEPARATES', 'OUTERWEAR', 'ACCESSORIES', 'FABRIC LINING', 'BIZ CARE PRINTED SCRUBS', 'PINK NBCF  |  With each PINK item sold, Biz Care will donate a percentage of proceeds to the National Breast Cancer Foundation.', 
                     'FUNCTIONAL COMFORT WAIST LOWERS - WOMENS', 'BIZ CARE AVERY SCRUBS', 'FLORENCE EASY STRETCH TOPS - DAISY PRINT', 'FUNCTIONAL STRETCH LOWERS - WOMENS', 
                     'BIZ CARE TOKYO SCRUBS', 'BIZ CARE RILEY SCRUBS', 'BIZ CARE PARKS SCRUBS', 'FUNCTIONAL COMFORT WAIST LOWERS - MENS', 'BIZ CARE HARTWELL SCRUBS', 'T-TOPS',
                     'BIZ CARE ROSE SCRUBS', 'PINK NBCF  |  With each PINK item sold, Biz Care will donate a percentage of proceeds to the Australian National Breast Cancer Foundation.',
                     'WORKS SHIRTS & POLOS', 'TTMC-W17 NEW ZEALAND', 'LAB COATS']
      
    # Make sure this is the correct column i.e. the unique identifier
    cell_product_slug = sh.cell(row=i, column=1).value
    if (cell_product_slug == None or (cell_product_slug in category_name)):
        continue
    print(cell_product_slug)
    pro_slug = str(cell_product_slug).strip()
    pro_slug = pro_slug.lower()

    price1 = sh.cell(row=i, column=6).value
    if price1 == None:
        continue
    price1 = str(price1)
    if '$' in price1:
        price1 = price1.replace('$', '')
    price1 = float(price1)
    # if isinstance(price1, str):
    #     price1 = price1.replace('$', '')
    #     price1 = float(price1)
    # price2 = sh.cell(row=i, column=8).value
    # price3 = sh.cell(row=i, column=9).value
    # price4 = sh.cell(row=i, column=10).value
    # price5 = sh.cell(row=i, column=11).value
    # price6 = sh.cell(row=i, column=12).value
    
    # This is so that there are no duplicates in the end list
    new_list_value = [pro_slug, price1]
    if new_list_value in list1:
        continue
    else:
        list1.append(new_list_value)

# Below code is there to quickly eliminate duplicates from list if not using the above method
# unique_list = list(dict.fromkeys(list1))
# print(unique_list)
print(list1)
